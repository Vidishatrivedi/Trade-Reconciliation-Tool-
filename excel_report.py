import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import date

# === LOAD DATA ===
breaks_df   = pd.read_csv("output/breaks_report.csv")
sig         = pd.read_csv("data/sig_ledger.csv")
cp          = pd.read_csv("data/counterparty_ledger.csv")

# Counterparty summary
cp_summary = (
    breaks_df.groupby("counterparty")["break_type"]
    .count()
    .reset_index()
    .rename(columns={"break_type": "total_breaks"})
    .sort_values("total_breaks", ascending=False)
    .reset_index(drop=True)
)

print("Data loaded")

# === DEFINE COLOURS AND STYLES ===

# Background colours
RED     = PatternFill("solid", fgColor="FF4444")
YELLOW  = PatternFill("solid", fgColor="FFD966")
GREEN   = PatternFill("solid", fgColor="70AD47")
BLUE    = PatternFill("solid", fgColor="1F4E79")
GREY    = PatternFill("solid", fgColor="F2F2F2")

# Font styles
WHITE_BOLD  = Font(color="FFFFFF", bold=True, size=11)
BLACK_BOLD  = Font(bold=True, size=11)
NORMAL      = Font(size=10)

# Alignment
CENTER  = Alignment(horizontal="center", vertical="center")
LEFT    = Alignment(horizontal="left",   vertical="center")

# Border
THIN_BORDER = Border(
    left=   Side(style="thin"),
    right=  Side(style="thin"),
    top=    Side(style="thin"),
    bottom= Side(style="thin")
)

print("Styles defined")

# === CREATE WORKBOOK ===
wb = openpyxl.Workbook()

# === SHEET 1 — SUMMARY DASHBOARD ===
ws1 = wb.active
ws1.title = "Summary"

# --- Title ---
ws1.merge_cells("A1:D1")
title_cell = ws1["A1"]
title_cell.value     = "TRADE RECONCILIATION REPORT"
title_cell.font      = WHITE_BOLD
title_cell.fill      = BLUE
title_cell.alignment = CENTER
ws1.row_dimensions[1].height = 30

# --- Date ---
ws1.merge_cells("A2:D2")
date_cell = ws1["A2"]
date_cell.value     = f"Generated: {date.today().strftime('%d %B %Y')}"
date_cell.font      = BLACK_BOLD
date_cell.alignment = CENTER
date_cell.fill      = GREY
ws1.row_dimensions[2].height = 20

# --- Summary Statistics ---
# Headers
stats = [
    ("METRIC",                  "VALUE",                ""),
    ("Total Trades Reviewed",   len(sig),               ""),
    ("Total Breaks Found",      len(breaks_df),         ""),
    ("Price Breaks",            len(breaks_df[breaks_df["break_type"] == "PRICE"]),    ""),
    ("Quantity Breaks",         len(breaks_df[breaks_df["break_type"] == "QUANTITY"]), ""),
    ("Status Breaks",           len(breaks_df[breaks_df["break_type"] == "STATUS"]),   ""),
    ("HIGH Severity Breaks",    len(breaks_df[breaks_df["severity"] == "HIGH"]),       ""),
    ("LOW Severity Breaks",     len(breaks_df[breaks_df["severity"] == "LOW"]),        ""),
    ("Break Rate",              f"{round(len(breaks_df)/len(sig)*100, 1)}%",           ""),
]

for row_num, (metric, value, _) in enumerate(stats, start=4):
    cell_a = ws1.cell(row=row_num, column=1, value=metric)
    cell_b = ws1.cell(row=row_num, column=2, value=value)

    cell_a.border    = THIN_BORDER
    cell_b.border    = THIN_BORDER
    cell_a.alignment = LEFT
    cell_b.alignment = CENTER

    if metric == "METRIC":
        cell_a.font = WHITE_BOLD
        cell_b.font = WHITE_BOLD
        cell_a.fill = BLUE
        cell_b.fill = BLUE
    elif metric == "HIGH Severity Breaks":
        cell_a.fill = RED
        cell_b.fill = RED
        cell_a.font = WHITE_BOLD
        cell_b.font = WHITE_BOLD
    else:
        cell_a.font = NORMAL
        cell_b.font = NORMAL

# Set column widths
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 20

print("Summary sheet created")

# === SHEET 2 — ALL BREAKS ===
ws2 = wb.create_sheet("All Breaks")

# --- Header row ---
headers = [
    "Trade ID", "Trade Date", "Counterparty",
    "Instrument", "Break Type", "SIG Value",
    "CP Value", "Difference", "Severity"
]

for col_num, header in enumerate(headers, start=1):
    cell = ws2.cell(row=1, column=col_num, value=header)
    cell.font      = WHITE_BOLD
    cell.fill      = BLUE
    cell.alignment = CENTER
    cell.border    = THIN_BORDER

# --- Data rows ---
for row_num, row in enumerate(breaks_df.itertuples(), start=2):
    values = [
        row.trade_id,
        row.trade_date,
        row.counterparty,
        row.instrument,
        row.break_type,
        row.sig_value,
        row.cp_value,
        row.difference,
        row.severity
    ]

    for col_num, value in enumerate(values, start=1):
        cell = ws2.cell(row=row_num, column=col_num, value=value)
        cell.border    = THIN_BORDER
        cell.alignment = CENTER
        cell.font      = NORMAL

        # Colour code by severity
        if row.severity == "HIGH":
            cell.fill = RED
            cell.font = Font(size=10, bold=True, color="FFFFFF")
        elif row.severity == "LOW":
            cell.fill = YELLOW

# Set column widths
col_widths = [12, 12, 16, 12, 12, 12, 12, 12, 10]
for i, width in enumerate(col_widths, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = width

# Freeze the header row so it stays visible when scrolling
ws2.freeze_panes = "A2"

print("Breaks sheet created")

# === SHEET 3 — COUNTERPARTY SUMMARY ===
ws3 = wb.create_sheet("Counterparty Summary")

# --- Title ---
ws3.merge_cells("A1:C1")
t = ws3["A1"]
t.value     = "BREAKS BY COUNTERPARTY"
t.font      = WHITE_BOLD
t.fill      = BLUE
t.alignment = CENTER
ws3.row_dimensions[1].height = 25

# --- Headers ---
cp_headers = ["Counterparty", "Total Breaks", "Risk Level"]
for col_num, header in enumerate(cp_headers, start=1):
    cell = ws3.cell(row=2, column=col_num, value=header)
    cell.font      = BLACK_BOLD
    cell.fill      = GREY
    cell.alignment = CENTER
    cell.border    = THIN_BORDER

# --- Data ---
max_breaks = cp_summary["total_breaks"].max()

for row_num, row in enumerate(cp_summary.itertuples(), start=3):

    # Assign risk level based on break count
    if row.total_breaks >= max_breaks * 0.7:
        risk  = "HIGH RISK"
        color = RED
        font  = Font(size=10, bold=True, color="FFFFFF")
    elif row.total_breaks >= max_breaks * 0.4:
        risk  = "MEDIUM RISK"
        color = YELLOW
        font  = Font(size=10, bold=False)
    else:
        risk  = "LOW RISK"
        color = GREEN
        font  = Font(size=10, color="FFFFFF")

    values = [row.counterparty, row.total_breaks, risk]
    for col_num, value in enumerate(values, start=1):
        cell = ws3.cell(row=row_num, column=col_num, value=value)
        cell.fill      = color
        cell.font      = font
        cell.alignment = CENTER
        cell.border    = THIN_BORDER

# Column widths
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 15
ws3.column_dimensions["C"].width = 15

print("Counterparty sheet created")

# === SAVE THE FILE ===
output_path = "output/reconciliation_report.xlsx"
wb.save(output_path)

print(f"\n Excel report saved to {output_path}")
print(f" Sheets created: Summary, All Breaks, Counterparty Summary")
print(f"\n Report complete — open reconciliation_report.xlsx to view!")